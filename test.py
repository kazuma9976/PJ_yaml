import yaml
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font


# YAMLファイルを読み取る関数
def load_yaml(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        data = yaml.safe_load(file)
    return data


# ネストしたデータをフラット化する関数
def flatten_dict(d, parent_key='', sep='.'):
    """辞書データをフラット化"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            items.append((new_key, ', '.join(map(str, v))))
        else:
            items.append((new_key, v))
    return dict(items)


# Excelファイルを作成する関数
def yaml_to_excel(yaml_data, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "yamlData"

    # スタイル設定
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    alignment = Alignment(horizontal='center', vertical='center')

    # データをフラット化して統一
    flattened_data = [flatten_dict(item) for item in yaml_data]
    headers = sorted(set(key for item in flattened_data for key in item.keys()))

    # ヘッダー行の設定
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = alignment

    # データ行の設定
    for row_num, item in enumerate(flattened_data, start=2):
        for col_num, header in enumerate(headers, start=1):
            value = item.get(header, "")
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = alignment

    # 列幅自動調整
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2

    # Excelファイルの保存
    wb.save(output_file)
    print(f"Excelファイルが保存されました: {output_file}")


# 実行例
yaml_file = "example.yaml"  # YAMLファイルのパス
excel_file = "output_nested.xlsx"  # 出力するExcelファイルのパス

yaml_data = load_yaml(yaml_file)
yaml_to_excel(yaml_data, excel_file)
